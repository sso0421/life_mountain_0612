import os
import random
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
import base64

# ==========================================
# 1. 페이지 기본 설정 및 디자인 테마
# ==========================================
st.set_page_config(
    page_title="산으로 가는 내인생",
    page_icon="🏔️",
    layout="centered"
)

# 자연적이고 편안한 그린(Forest Green) 계열 및 게이지 바 / 링크 제거 스타일링
st.markdown("""
    <style>
    /* 전체 앱 배경색 지정 */
    .stApp {
        background-color: #F9FBE7 !important;
    }
    
    /* 인트로 전용 큰 아이콘 스타일 */
    .intro-icon { 
        font-size: 5.5rem; 
        text-align: center; 
        margin-top: 20px;
        margin-bottom: 10px; 
    }
    
    .main-title { font-size: 2.6rem; font-weight: 800; color: #2C5E3B; text-align: center; margin-bottom: 5px; }
    .sub-title { font-size: 1.2rem; color: #4E7055; text-align: center; margin-bottom: 25px; font-weight: 500; }
    .desc-box { background-color: #F1F6F2; padding: 20px; border-radius: 10px; border-left: 5px solid #2C5E3B; margin-bottom: 30px; }
    
    /* 하단 이동 버튼 전용 스타일: 흰색 배경 + 초록 테두리 + 초록 글자 */
    .stButton>button { 
        background-color: #FFFFFF !important; 
        color: #2C5E3B !important; 
        border: 2px solid #2C5E3B !important;
        border-radius: 6px; 
        width: 100%; 
        font-size: 16px; 
        font-weight: bold; 
        padding: 10px;
        transition: all 0.3s ease;
    }
    .stButton>button:hover { 
        background-color: #F1F6F2 !important; 
        color: #1E4228 !important;
        border-color: #1E4228 !important;
    }
    
    /* 게이지 바 색상 반전 처리 */
    div[data-baseweb="progress-bar"] > div > div {
        background-color: #2C5E3B !important;
    }
    div[data-testid="stProgress"] div div div div {
        background-color: #2C5E3B !important;
    }
    div[data-baseweb="progress-bar"] {
        background-color: #FFFFFF !important;
        border: 1px solid #E0E0E0;
        border-radius: 4px;
    }
    div[data-testid="stProgress"] > div {
        background-color: #FFFFFF !important;
    }
    
    /* 제목 영역 마우스 호버 시 나타나는 링크 아이콘(Anchor link) 제거 */
    a.anchor-link {
        display: none !important;
    }
    .stMarkdown h1 a, .stMarkdown h2 a, .stMarkdown h3 a, .stMarkdown h4 a {
        display: none !important;
    }
    
    /* 기운 요약 카드 스타일 */
    .energy-card {
        text-align: center;
        padding: 15px 5px;
        background-color: #FAFAFA;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    .energy-icon { font-size: 2.2rem; margin-bottom: 5px; }
    .energy-name { font-size: 1rem; font-weight: bold; color: #555555; margin-bottom: 8px; }
    .energy-value { font-size: 1.8rem; font-weight: 800; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 2. 데이터 로드 및 세션 상태 초기화
# ==========================================
@st.cache_data
def load_db_files():
    q_df = pd.read_excel("question_DB.xlsx")
    a_df = pd.read_excel("answer_DB.xlsx")
    return q_df, a_df

try:
    q_df, a_df = load_db_files()
except Exception as e:
    st.error(f"엑셀 파일을 불러오는 중 오류가 발생했습니다. 파일명과 위치를 확인해 주세요. ({e})")
    st.stop()

# 규칙 및 매핑 정의
CATEGORY_MAPPING = {
    "선택항목1": "재물",
    "선택항목2": "건강",
    "선택항목3": "행복",
    "선택항목4": "성공",
    "선택항목5": "애정"
}
PRIORITY_ORDER = ["재물", "건강", "행복", "성공", "애정"]

# 세션 관리 변수 초기화
if "page" not in st.session_state:
    st.session_state.page = "intro"
if "info_data" not in st.session_state:
    st.session_state.info_data = {}
if "test_answers" not in st.session_state:
    st.session_state.test_answers = {}  
if "test_selected_texts" not in st.session_state:
    st.session_state.test_selected_texts = {} 
if "current_q_idx" not in st.session_state:
    st.session_state.current_q_idx = 0

# ==========================================
# 3. 데이터 저장 및 파일 처리 함수
# ==========================================
def save_to_result_db(final_row):
    file_name = "result_DB.xlsx"
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        if ws.max_row == 1:
            next_no = 1
        else:
            last_val = ws.cell(row=ws.max_row, column=1).value
            try:
                next_no = int(last_val) + 1
            except:
                next_no = ws.max_row
    else:
        wb = Workbook()
        ws = wb.active
        headers = [
            "no", "별명_이름", "성별", "연령", "지역", 
            "테스트1", "테스트2", "테스트3", "테스트4", "테스트5", 
            "테스트6", "테스트7", "테스트8", "테스트9", "테스트결과값", "추천산"
        ]
        ws.append(headers)
        next_no = 1
        
    full_row = [next_no] + final_row
    ws.append(full_row)
    wb.save(file_name)

def find_image_path(region, mountain_name):
    base_dir = os.path.join("mountain", str(region))
    if os.path.exists(base_dir):
        for ext in ['.png', '.jpg', '.jpeg', '.PNG', '.JPG']:
            img_path = os.path.join(base_dir, f"{mountain_name}{ext}")
            if os.path.exists(img_path):
                return img_path
    return None

# [디자인/기능 전면 개정] 요청된 고유 색상 매핑이 적용된 다중 마우스 오버 툴팁 시스템
def render_donut_chart(df, region_filter=None):
    categories = ["재물", "건강", "행복", "성공", "애정"]
    colors = ["#F7D200", "#8600AF", "#007C06", "#055FBE", "#FF699B"] # 노랑, 보라, 초록, 파랑, 분홍
    
    if df.empty or "테스트결과값" not in df.columns:
        return "<div style='text-align: center; color: #888888; padding: 40px 0; font-family: sans-serif; font-weight: 500;'>수집된 데이터가 없습니다.</div>"
        
    if region_filter:
        if "지역" in df.columns:
            filtered_df = df[df["지역"] == region_filter]
        else:
            filtered_df = pd.DataFrame()
    else:
        filtered_df = df
        
    if filtered_df.empty:
        return "<div style='text-align: center; color: #888888; padding: 40px 0; font-family: sans-serif; font-weight: 500;'>수집된 데이터가 없습니다.</div>"
        
    counts = filtered_df["테스트결과값"].value_counts().to_dict()
    total = sum(counts.values())
    
    if total == 0:
        return "<div style='text-align: center; color: #888888; padding: 40px 0; font-family: sans-serif; font-weight: 500;'>수집된 데이터가 없습니다.</div>"
        
    gradient_parts = []
    current_pct = 0.0
    hover_title_parts = [] 
    
    for cat, color in zip(categories, colors):
        cnt = counts.get(cat, 0)
        if cnt > 0:
            pct = (cnt / total) * 100
            next_pct = current_pct + pct
            gradient_parts.append(f"{color} {current_pct:.1f}% {next_pct:.1f}%")
            current_pct = next_pct
            hover_title_parts.append(f"{cat}: {cnt}명({pct:.0f}%)")
            
    gradient_str = ", ".join(gradient_parts)
    donut_hover_title = " / ".join(hover_title_parts) 
    
    if not gradient_parts:
        return "<div style='text-align: center; color: #888888; padding: 40px 0; font-family: sans-serif; font-weight: 500;'>수집된 데이터가 없습니다.</div>"
        
    legend_items_html = ""
    for cat, color in zip(categories, colors):
        cnt = counts.get(cat, 0)
        pct = (cnt / total) * 100 if total > 0 else 0
        
        legend_items_html += f"""
        <div style="display: flex; align-items: center; justify-content: center; background: #FFFFFF; border: 1px solid #EAEAEA; border-radius: 6px; padding: 6px 4px; box-sizing: border-box; transition: all 0.2s ease; cursor: pointer;" title="선택한 누적 인원: {cnt}명">
            <span style="display: inline-block; width: 9px; height: 9px; background-color: {color}; border-radius: 50%; margin-right: 6px; flex-shrink: 0;"></span>
            <span style="color: #555555; font-size: 0.8rem; font-weight: 600; margin-right: 4px; white-space: nowrap;">{cat}</span>
            <span style="font-weight: 700; color: {color}; font-size: 0.8rem; white-space: nowrap;">{pct:.0f}%</span>
        </div>
        """
        
    html_str = f"""
    <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; font-family: 'Malgun Gothic', sans-serif; background-color: #FAFAFA; padding: 18px; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.02); box-sizing: border-box; width: 100%;">
        
        <div style="background: conic-gradient({gradient_str}); border-radius: 50%; width: 130px; height: 130px; display: flex; align-items: center; justify-content: center; box-shadow: inset 0 0 6px rgba(0,0,0,0.04); cursor: pointer;" title="{donut_hover_title}">
            <div style="background: #FAFAFA; border-radius: 50%; width: 88px; height: 88px; display: flex; flex-direction: column; align-items: center; justify-content: center;">
                <span style="font-size: 0.7rem; color: #999999; margin-bottom: 1px; letter-spacing: -0.5px;">누적 참여</span>
                <span style="font-weight: 800; color: #2C5E3B; font-size: 1.1rem;">{total}명</span>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 6px; width: 100%; margin-top: 18px; box-sizing: border-box;">
            {legend_items_html}
        </div>
    </div>
    """
    return html_str

# ==========================================
# 4. 앱 화면 및 비즈니스 로직 레이아웃
# ==========================================

# --- [1] 인트로 화면 ---
if st.session_state.page == "intro":
    
    # 1. 로컬 이미지 파일 읽어서 Base64 인코딩
    img_filename = "image.png" 
    try:
        with open(img_filename, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()
        
        # 2. 블러 및 흰색 필터가 적용된 CSS 주입
        st.markdown(f"""
            <style>
            /* 배경 이미지와 필터를 적용할 가상 레이어 생성 */
            .stApp::before {{
                content: "";
                position: fixed;
                top: 0;
                left: 0;
                width: 100%;
                height: 100%;
                
                /* 배경 이미지 설정 */
                background-image: url("data:image/jpeg;base64,{encoded_string}");
                background-size: cover;
                background-position: center;
                background-repeat: no-repeat;
                
                /* 블러 처리 */
                filter: blur(5px);
                
                /* 흰색 느낌 추가 */
                background-color: rgba(255, 255, 255, 0.45);
                background-blend-mode: overlay;
                
                /* 콘텐츠 뒤로 보내기 */
                z-index: -1;
                transform: scale(1.05);
            }}
            
            /* 전체 앱 본문 배경은 투명하게 처리하여 가상 레이어가 보이도록 설정 */
            .stApp {{
                background-color: transparent !important;
            }}

            /* 타이틀 및 서브타이틀 시인성 조정 */
            .main-title {{
                color: #1E4228 !important;
                font-weight: 800;
            }}
            .sub-title {{
                color: #3A5F43 !important;
                font-weight: 600;
            }}
            
            /* 설명 박스는 배경이 밝아졌으므로 약간 더 선명하게 조정 */
            .desc-box {{
                background-color: rgba(255, 255, 255, 0.9) !important;
                box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
            }}
            </style>
        """, unsafe_allow_html=True)
    except FileNotFoundError:
        pass

    # 3. 컴포넌트 배치
    st.markdown('<div class="main-title">🏔️산으로 가는 내인생</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">나의 부족한 기운을 채워줄, 개인 맞춤형 명산 큐레이터!</div>', unsafe_allow_html=True)

    st.markdown('<div class="intro-icon">🌲 🏔️ 🎒 ⛺</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="desc-box">
        <h4 style="margin-top:0; color:#2C5E3B;">📋 테스트 설명</h4>
        <p style="margin:0; font-size:1.05rem; line-height:1.6; color:#2F4F4F;">
            당신에게 필요한 기운을 채워 줄 <b>산 추천</b>과 더불어, <br>
            답답한 일상을 리프레시할 <b>등산 코스, 주변 맛집, 행운의 아이템 추천</b>까지 한 번에 만나보세요!
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("테스트 시작하기", use_container_width=True):
        st.session_state.page = "info"
        st.rerun()

# --- [2] 개인정보 및 지역 설정 영역 ---
elif st.session_state.page == "info":
    st.markdown('<div class="main-title">🏔️ 산으로 가는 내인생</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">나의 부족한 기운을 채워줄, 개인 맞춤형 명산 큐레이터!</div>', unsafe_allow_html=True)
    
    st.subheader("📋 기본 정보 및 지역 선택")
    st.caption("모든 문항은 필수 입력 사항입니다. (디폴트 선택 없음)")
    
    saved_info = st.session_state.info_data
    default_name = saved_info.get("name", "")
    
    q1_row = q_df[q_df['질문번호'] == 1].iloc[0]
    name_input = st.text_input(f"Q1. {q1_row['질문내용']}", value=default_name, key="info_name", placeholder="별명이나 이름을 입력해 주세요.")
    
    q2_row = q_df[q_df['질문번호'] == 2].iloc[0]
    q2_opts = [q2_row['선택항목1'], q2_row['선택항목2'], q2_row['선택항목3']]
    default_gender_idx = q2_opts.index(saved_info["gender"]) if "gender" in saved_info and saved_info["gender"] in q2_opts else None
    gender_input = st.radio(f"Q2. {q2_row['질문내용']}", options=q2_opts, index=default_gender_idx, key="info_gender")
    
    q3_row = q_df[q_df['질문번호'] == 3].iloc[0]
    q3_opts = [q3_row[f'선택항목{i}'] for i in range(1, 10) if pd.notna(q3_row[f'선택항목{i}']) and str(q3_row[f'선택항목{i}']).strip() != ""]
    default_age_idx = q3_opts.index(saved_info["age"]) if "age" in saved_info and saved_info["age"] in q3_opts else None
    age_input = st.radio(f"Q3. {q3_row['질문내용']}", options=q3_opts, index=default_age_idx, key="info_age")
    
    q4_row = q_df[q_df['질문번호'] == 4].iloc[0]
    q4_opts = [q4_row[f'선택항목{i}'] for i in range(1, 10) if pd.notna(q4_row[f'선택항목{i}']) and str(q4_row[f'선택항목{i}']).strip() != ""]
    default_region_idx = q4_opts.index(saved_info["region"]) if "region" in saved_info and saved_info["region"] in q4_opts else None
    region_input = st.radio(f"Q4. {q4_row['질문내용']}", options=q4_opts, index=default_region_idx, key="info_region")
    
    st.markdown("---")
    if st.button("다음 단계로 이동 >"):
        if not name_input.strip():
            st.warning("이름 또는 별명을 꼭 입력해 주세요.")
        elif gender_input is None:
            st.warning("성별을 선택해 주세요.")
        elif age_input is None:
            st.warning("연령대를 선택해 주세요.")
        elif region_input is None:
            st.warning("원하는 지역을 선택해 주세요.")
        else:
            st.session_state.info_data = {
                "name": name_input.strip(),
                "gender": gender_input,
                "age": age_input,
                "region": region_input
            }
            st.session_state.page = "test"
            st.session_state.current_q_idx = 0
            st.rerun()

# --- [3] 테스트 문항 개별 순차 노출 영역 ---
elif st.session_state.page == "test":
    st.markdown('<div class="main-title">🏔️ 산으로 가는 내인생</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">나의 부족한 기운을 채워줄, 개인 맞춤형 명산 큐레이터!</div>', unsafe_allow_html=True)
    
    test_questions = q_df[(q_df['질문번호'] >= 5) & (q_df['질문번호'] <= 13)].reset_index(drop=True)
    total_q_count = len(test_questions)
    current_idx = st.session_state.current_q_idx
    
    if current_idx < total_q_count:
        row = test_questions.iloc[current_idx]
        q_num = int(row['질문번호'])
        display_q_num = current_idx + 1
        
        if 1 <= display_q_num <= 4:
            progress_icon = "🧗‍♂️ 산을 오르는 중..."
        elif 5 <= display_q_num <= 6:
            progress_icon = "🏔️ 정상에 도착!"
        elif 7 <= display_q_num <= 8:
            progress_icon = "🚶‍♂️ 하산하는 중..."
        elif display_q_num == 9:
            progress_icon = "⛺ 꿀맛 같은 휴식"
            
        passed_path = "🚩" * current_idx
        remaining_path = "🏔️" * (total_q_count - current_idx - 1)
        
        st.markdown(f"""
        <div style='text-align: center; margin-bottom: 25px;'>
            <div style='font-size: 1.1rem; color: #2C5E3B; font-weight: bold;'>진행도: {display_q_num} / {total_q_count}</div>
            <div style='font-size: 0.95rem; color: #555555; margin-top: 3px; margin-bottom: 8px; font-weight: 500;'>현재 상태: {progress_icon}</div>
            <div style='font-size: 2.2rem; letter-spacing: 2px;'>{passed_path}🚶‍♂️{remaining_path}</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.subheader(f"Q{display_q_num}. {row['질문내용']}")
        
        options_map = {}
        for i in range(1, 10):
            col_name = f'선택항목{i}'
            if pd.notna(row[col_name]) and str(row[col_name]).strip() != "":
                text = str(row[col_name])
                energy_type = CATEGORY_MAPPING.get(col_name, "기타")
                options_map[text] = energy_type
        
        # [수정] 질문별로 선택 답안의 순서를 랜덤하게 섞음 (페이지 Rerun 시 순서가 유지되도록 세션 상태 사용)
        shuffle_key = f"shuffled_options_{q_num}"
        if shuffle_key not in st.session_state:
            raw_options = list(options_map.keys())
            random.shuffle(raw_options)
            st.session_state[shuffle_key] = raw_options
            
        options_list = st.session_state[shuffle_key]
        
        default_index = None
        if q_num in st.session_state.test_selected_texts:
            saved_text = st.session_state.test_selected_texts[q_num]
            if saved_text in options_list:
                default_index = options_list.index(saved_text)
                
        selected_text = st.radio(
            "마음에 드는 항목을 하나 선택하세요.",
            options=options_list,
            index=default_index,
            key=f"test_q_{q_num}"
        )
        
        st.markdown("---")
        
        is_last = (display_q_num == total_q_count)
        col_prev, col_spacer, col_next = st.columns([1, 1.5, 1])  
        
        with col_prev:
            if current_idx == 0:
                if st.button("<< 처음으로"):
                    st.session_state.page = "info"
                    st.rerun()
            elif current_idx > 0:
                if st.button("< 이전 문항"):
                    st.session_state.current_q_idx -= 1
                    st.rerun()
                    
        with col_next:
            btn_label = "결과 분석하기" if is_last else "다음 문항으로 >"
            
            if st.button(btn_label):
                if selected_text is None:
                    st.warning("답변을 선택하셔야 다음 단계 진행이 가능합니다.")
                else:
                    st.session_state.test_answers[q_num] = options_map[selected_text]
                    st.session_state.test_selected_texts[q_num] = selected_text 
                    st.session_state.current_q_idx += 1
                    
                    if is_last:
                        st.session_state.page = "result"
                    st.rerun()

# --- [4] 결과 도출 및 대시보드 리포트 영역 ---
elif st.session_state.page == "result":
    st.markdown('<div class="main-title">🏔️ 산으로 가는 내인생</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">나의 부족한 기운을 채워줄, 개인 맞춤형 명산 큐레이터!</div>', unsafe_allow_html=True)
    
    info = st.session_state.info_data
    answers = st.session_state.test_answers
    
    chosen_energies = list(answers.values())
    max_count = -1
    final_result_value = None
    
    for energy in PRIORITY_ORDER:
        cnt = chosen_energies.count(energy)
        if cnt > max_count:
            max_count = cnt
            final_result_value = energy
            
    matched_data = a_df[(a_df['지역'] == info['region']) & (a_df['카테고리'] == final_result_value)]
    
    if not matched_data.empty:
        rec_row = matched_data.iloc[0]
    else:
        fallback_data = a_df[a_df['지역'] == info['region']]
        rec_row = fallback_data.iloc[0] if not fallback_data.empty else a_df.iloc[0]
        
    rec_mountain = rec_row['추천산']
    
    if "db_saved" not in st.session_state:
        test_fields = [answers[i] for i in range(5, 14)]
        final_row_structure = [
            info['name'], info['gender'], info['age'], info['region']
        ] + test_fields + [final_result_value, rec_mountain]
        
        try:
            save_to_result_db(final_row_structure)
            st.session_state.db_saved = True
        except Exception as e:
            st.error(f"결과 데이터베이스를 업데이트하는 중 오류가 발생했습니다: {e}")

    st.balloons()
    
    st.markdown(f"""
    <div style="text-align: center; margin-top: 10px; margin-bottom: 25px;">
        <span style="background-color: #2C5E3B; color: white; padding: 6px 16px; border-radius: 20px; font-size: 0.9rem; font-weight: bold;">ANALYTIC REPORT</span>
        <h2 style="color: #1E4228; margin-top: 12px; line-height: 1.5;">
            <b>{info['name']}</b>님의 부족한 기운을 채워줄 기운은<br><b>'{final_result_value}'</b>입니다!
        </h2>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 나의 5가지 균형 점수")
    
    card_config = {
        "건강": {"icon": "💪", "color": "#8600AFFF"},
        "행복": {"icon": "🍀", "color": "#007C06FF"},
        "애정": {"icon": "❤️", "color": "#FF699BFF"},
        "재물": {"icon": "💎", "color": "#F7D200"},
        "성공": {"icon": "🏆", "color": "#055FBE"}
    }
    
    total_questions_count = len(chosen_energies) 
    
    calculated_energies = []
    for energy_name in ["건강", "행복", "애정", "재물", "성공"]:
        count = chosen_energies.count(energy_name)
        percentage = round((count / total_questions_count) * 100) if total_questions_count > 0 else 0
        calculated_energies.append({
            "name": energy_name,
            "percentage": percentage,
            "config": card_config[energy_name]
        })
        
    sorted_energies = sorted(calculated_energies, key=lambda x: x['percentage'], reverse=True)
    
    card_cols = st.columns(5)
    for idx, item in enumerate(sorted_energies):
        with card_cols[idx]:
            st.markdown(f"""
                <div class="energy-card">
                    <div class="energy-icon">{item['config']['icon']}</div>
                    <div class="energy-name">{item['name']}</div>
                    <div class="energy-value" style="color: {item['config']['color']};">{item['percentage']}%</div>
                </div>
            """, unsafe_allow_html=True)
            
    st.markdown("---")
    
    st.markdown("### 다른 유저들은 어떤 기운이 부족할까요?")
    
    db_df = pd.DataFrame()
    if os.path.exists("result_DB.xlsx"):
        try:
            db_df = pd.read_excel("result_DB.xlsx")
        except:
            pass
            
    col_graph1, col_graph2 = st.columns(2)
    with col_graph1:
        st.markdown("<h4 style='text-align: center; color: #2C5E3B; margin-bottom: 5px;'>🌍 전체 지역 기준</h4>", unsafe_allow_html=True)
        global_donut_html = render_donut_chart(db_df, region_filter=None)
        st.components.v1.html(global_donut_html, height=250, scrolling=False)
        
    with col_graph2:
        st.markdown(f"<h4 style='text-align: center; color: #2C5E3B; margin-bottom: 5px;'>📍 {info['region']} 지역 기준</h4>", unsafe_allow_html=True)
        region_donut_html = render_donut_chart(db_df, region_filter=info['region'])
        st.components.v1.html(region_donut_html, height=250, scrolling=False)
        
    st.markdown("---")
    
    col_visual, col_content = st.columns([1, 1.2])
    
    with col_visual:
        img_file_path = find_image_path(info['region'], rec_mountain)
        if img_file_path:
            st.image(img_file_path, use_container_width=True, caption=f"{info['region']} {rec_mountain}")
        else:
            st.info(f"[이미지 준비 중]\nmountain/{info['region']}/{rec_mountain}")

    with col_content:
        st.markdown(f"### 추천 명산: **{rec_mountain}**")
        if '위치_고도' in rec_row and pd.notna(rec_row['위치_고도']):
            st.caption(f"위치/고도: {rec_row['위치_고도']}")
            
        st.markdown("**💡 산의 기운**")
        st.write(rec_row['카테고리설명'])
        
        st.markdown("**📜 명산에 얽힌 스토리**")
        st.write(rec_row['스토리'])

    st.markdown("---")
    # 추천 코스 및 거리/시간 통합 가이드 박스 레이아웃
    # [디자인 전면 개정] 대시보드 테마와 조화를 이루는 연초록 글래스모피즘 가이드 카드
    st.markdown(f"""
    <div style="background-color: rgba(241, 246, 242, 0.85); border-left: 6px solid #2C5E3B; border-radius: 10px; padding: 22px 26px; box-shadow: 0 4px 12px rgba(44, 94, 59, 0.05); margin-bottom: 25px;">
        <h3 style="margin-top: 0; color: #2C5E3B; font-size: 1.35rem; font-weight: 800; margin-bottom: 18px; display: flex; align-items: center;">
            <span style="margin-right: 8px; font-size: 1.5rem;">🥾</span> 추천 등산 코스
        </h3>
        <div style="margin-bottom: 12px; line-height: 1.6; color: #2F4F4F; font-size: 1.05rem;">
            <strong style="color: #2C5E3B;">코스 :</strong> 
            <span style="margin-left: 6px; font-weight: 500;">{rec_row['추천코스']}</span>
        </div>
        <div style="line-height: 1.6; color: #2F4F4F; font-size: 1.05rem;">
            <strong style="color: #2C5E3B;">거리 및 시간 :</strong> 
            <span style="margin-left: 6px; font-weight: 500;">{rec_row['코스왕복거리_시간']}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
        
    raw_items = [rec_row['행운아이템1'], rec_row['행운아이템2'], rec_row['행운아이템3']]
    clean_items = [it for it in raw_items if pd.notna(it) and str(it).strip() != ""]
    selected_lucky_item = random.choice(clean_items) if clean_items else "나만의 등산 지팡이"
    
    col_rest, col_lucky = st.columns(2)
    with col_rest:
        st.markdown("🍕 **산악 매칭 추천 맛집**")
        st.warning(f"**{rec_row['맛집']}**")
        
    with col_lucky:
        st.markdown("🍀 **오늘의 행운의 아이템**")
        st.info(f"**{selected_lucky_item}**")
    
    st.markdown("---")
    if st.button("테스트 다시 참여하기"):
        for k in ["info_data", "test_answers", "test_selected_texts", "current_q_idx", "db_saved"]:
            if k in st.session_state:
                del st.session_state[k]
        # [수정] 테스트 다시 참여 시 셔플된 선택지 목록도 초기화
        keys_to_del = [k for k in st.session_state.keys() if k.startswith("shuffled_options_")]
        for k in keys_to_del:
            del st.session_state[k]
        st.session_state.page = "intro"
        st.rerun()
